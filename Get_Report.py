from datetime import datetime
from tkinter import Tk,filedialog
from configparser import ConfigParser
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment,Font,PatternFill
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00,FORMAT_NUMBER_00
from configparser import ConfigParser
from glob import glob
import os,re,sys

root=Tk()
root.withdraw()

# filediag=filedialog.askopenfile(initialdir='.',title="Choose one log directory",filetypes=[('dir','*.ini')])
LogDir=filedialog.askdirectory(initialdir='.',title="Choose one log directory")

if LogDir == "":
    sys.exit()

    
src_file=glob(LogDir+"\*\*.ini")
src_file_list = [x for x in src_file if re.search(".*\\\\(TAIWIS|CHOWIS)\w*\.ini",x)]
servers=[]
for x in glob(LogDir+"\*\*.ini"):
    servers.append((os.path.basename(x)).replace(".ini",""))
server_list=[]
for x in servers:
    if re.search("^(TAIWIS|CHOWIS)\w*",x):
        server_list.append(x)
server_list.reverse()
src_file_list.reverse()

config=ConfigParser(interpolation=None)
Wb=Workbook()
Ws=Wb.active
header=["Hostname","Free C: %","Free D: %","Free E: %","LastBootTime","LastDefragTime",\
        "LastAVUpdate","LastScanTime","LastWindowsUpdate",\
        "ATF","LAPS","DPS","WDT","PCR","PoConfig"]
Ws.append(header)

Ws.column_dimensions[get_column_letter(1)].width=21
for row in Ws[1]:
    row.alignment=Alignment(horizontal="center",vertical="center")
for r in range(1,len(server_list)+1):
    #print(server_list[r-1])
    Ws.cell(r+1,1).value=server_list[r-1]
    Ws.cell(r+1,1).font=Font(sz=12)
    Ws.row_dimensions[r].height=20
    #Ws.column_dimensions[get_column_letter(r)].width=20
    Ws.cell(r+1,1).alignment=Alignment(vertical="center")
    
row=2
#print(len(header))
for file in src_file_list:
    config.read(file)
    print(file)
    for col,value in enumerate(config["Disks Free Rate"].values(),2):
        Ws.cell(row,col).value=float(value)/100
        Ws.cell(row,col).font=Font(sz=12)
        Ws.cell(row,col).number_format=FORMAT_PERCENTAGE_00
        Ws.cell(row,col).alignment=Alignment(horizontal="right")

        if float(value)/100 <= 0.15 :
            Ws.cell(row,col).fill=PatternFill(fill_type="solid",fgColor="00FFFF00")
            Ws.cell(row,1).font=Font(sz=12,bold=True)

    for col in range(5,len(header)+1):
        #print(header[col-1])
        if header[col-1] in config.sections() and Ws.cell(1,col).value == header[col-1] and Ws.cell(row,1).value == re.search("\w+(\.ini)",file).group().replace(".ini",""):
            Ws.cell(row,col).value=config.get(header[col-1],header[col-1])
            Ws.cell(row,col).font=Font(sz=12)
            Ws.cell(row,col).alignment=Alignment(vertical="center",horizontal="right")
            Ws.column_dimensions[get_column_letter(col)].width=len(Ws.cell(row,col).value)+4
        #print(Ws.cell(row,col).value)
        if Ws.cell(row,col).value is not None and re.search("^\d{4}\-\d{2}-\d{2}\s\d{2}\:\d{2}\:\d{2}",Ws.cell(row,col).value) :
            if (datetime.strptime(LogDir[-8:],"%Y%m%d")-datetime.strptime(Ws.cell(row,col).value,"%Y-%m-%d %H:%M:%S")).days >= 25:
                Ws.cell(row,col).fill=PatternFill(fill_type="solid",fgColor="00FFFF00")
                Ws.cell(row,1).font=Font(sz=12,bold=True)
            pass 
    config.clear()
    row+=1

for col in range(2,5):
    Ws.column_dimensions[get_column_letter(col)].width=len(Ws.cell(1,col).value)+3

# now=datetime.now().strftime("%Y%m%d")
Ws.freeze_panes = 'A2'
Wb.save("PMcheck_Report_"+LogDir[-8:]+".xlsx")
